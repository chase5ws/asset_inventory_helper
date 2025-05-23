from tkinter import Tk, ttk, filedialog, messagebox, StringVar, simpledialog
from openpyxl import load_workbook

def select_source_excel():
    file_path = filedialog.askopenfilename(
        title="選擇來源 Excel 文件",
        filetypes=[("Excel 文件", "*.xlsx *.xls")]
    )
    if file_path:
        source_excel_var.set(file_path)
        load_fields_and_dropdowns(file_path)

def select_target_excel():
    file_path = filedialog.askopenfilename(
        title="選擇目標 Excel 文件",
        filetypes=[("Excel 文件", "*.xlsx *.xls")]
    )
    if file_path:
        target_excel_var.set(file_path)
        load_sheets(file_path)

def load_fields_and_dropdowns(file_path):
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]  # 假設第一行是欄位名稱

        # 清空舊的欄位和控件
        for widget in frame_fields.winfo_children():
            widget.destroy()

        dropdowns.clear()

        # 動態生成控件
        for idx, field in enumerate(header):
            row = idx // 3
            col = (idx % 3) * 3

            ttk.Label(frame_fields, text=field).grid(row=row, column=col, padx=10, pady=10, sticky="w")
            dropdown = ttk.Combobox(frame_fields, width=20, state="readonly")
            dropdown.grid(row=row, column=col + 1, padx=10, pady=10)
            dropdowns[field] = dropdown

            manual_input_button = ttk.Button(frame_fields, text="手動輸入",
                                             command=lambda f=field: add_manual_input(f))
            manual_input_button.grid(row=row, column=col + 2, padx=10, pady=10)

        # 加載數據到下拉選單
        for field, dropdown in dropdowns.items():
            if field in header:
                col_index = header.index(field) + 1
                values = [sheet.cell(row=row, column=col_index).value for row in range(2, sheet.max_row + 1)]
                values = list(filter(None, values))
                dropdown["values"] = values
                if values:
                    dropdown.set(values[0])

    except Exception as e:
        messagebox.showerror("錯誤", f"讀取來源 Excel 文件時發生錯誤：{e}")

def load_sheets(file_path):
    try:
        workbook = load_workbook(file_path)
        sheet_names = workbook.sheetnames
        sheet_dropdown["values"] = sheet_names
        if sheet_names:
            sheet_dropdown.set(sheet_names[0])
    except Exception as e:
        messagebox.showerror("錯誤", f"讀取目標 Excel 文件時發生錯誤：{e}")

def add_manual_input(field):
    new_value = simpledialog.askstring("手動輸入", f"請輸入 {field} 的新值：")
    if new_value:
        dropdown = dropdowns[field]
        current_values = list(dropdown["values"])
        if new_value not in current_values:
            current_values.append(new_value)
            dropdown["values"] = current_values
        dropdown.set(new_value)

def add_data_to_target_excel():
    target_file = target_excel_var.get()
    selected_sheet = sheet_dropdown.get()

    if not target_file or not selected_sheet:
        messagebox.showerror("錯誤", "請選擇目標 Excel 文件和工作表！")
        return

    try:
        workbook = load_workbook(target_file)
        if selected_sheet not in workbook.sheetnames:
            messagebox.showerror("錯誤", f"工作表 {selected_sheet} 不存在於目標文件中！")
            return

        sheet = workbook[selected_sheet]
        last_row = sheet.max_row + 1

        for col, field in enumerate(dropdowns.keys(), start=1):
            sheet.cell(row=last_row, column=col).value = dropdowns[field].get()

        workbook.save(target_file)
        messagebox.showinfo("成功", "數據已成功新增到目標 Excel 文件！")
    except Exception as e:
        messagebox.showerror("錯誤", f"新增數據時發生錯誤：{e}")

def close_program():
    root.destroy()

if __name__ == "__main__":
    root = Tk()
    root.title("Excel 動態欄位管理系統")

    source_excel_var = StringVar()
    target_excel_var = StringVar()
    dropdowns = {}

    # 主界面設置
    ttk.Label(root, text="來源 Excel 文件：").grid(row=0, column=0, padx=10, pady=5, sticky="w")
    ttk.Entry(root, textvariable=source_excel_var, width=50, state="readonly").grid(row=0, column=1, columnspan=2, padx=10, pady=5)
    ttk.Button(root, text="選擇來源文件", command=select_source_excel).grid(row=0, column=3, padx=10, pady=5)

    ttk.Label(root, text="目標 Excel 文件：").grid(row=1, column=0, padx=10, pady=5, sticky="w")
    ttk.Entry(root, textvariable=target_excel_var, width=50, state="readonly").grid(row=1, column=1, columnspan=2, padx=10, pady=5)
    ttk.Button(root, text="選擇目標文件", command=select_target_excel).grid(row=1, column=3, padx=10, pady=5)

    ttk.Label(root, text="選擇工作表：").grid(row=2, column=0, padx=10, pady=5, sticky="w")
    sheet_dropdown = ttk.Combobox(root, width=47, state="readonly")
    sheet_dropdown.grid(row=2, column=1, columnspan=2, padx=10, pady=5)

    # 動態欄位框架
    frame_fields = ttk.Frame(root)
    frame_fields.grid(row=3, column=0, columnspan=4, padx=10, pady=10)

    ttk.Button(root, text="新增資料", command=add_data_to_target_excel).grid(row=4, column=0, columnspan=2, pady=10)
    ttk.Button(root, text="關閉程式", command=close_program).grid(row=4, column=2, columnspan=2, pady=10)

    root.mainloop()